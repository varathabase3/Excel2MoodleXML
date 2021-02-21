#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
@author: varatharajan
License: GNU GPL V3
"""


import string
import sys
# from datetime import datetime
from xml.dom import minidom
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import base64
from os import path
# import os
import logging
from io import BytesIO  # Saving image files to buffer before converting to BASE64


def check_and_break_lines_with_br(txt):
    qn_stack = 0  # This is used to create a virtual stack index to track the presence of equations
    no_of_question = 0
    # index = txt.find('\(')

    # lines = txt.splitlines()
    txt = txt.replace('\n', ' <br> ')  # replace next line with HTML <br> (break) code
    str_array = txt.split()
    for word in str_array:
        if word.startswith('\\('):
            qn_stack = qn_stack + 1
        if word.__contains__('\\)'):
            if qn_stack > 0:
                qn_stack = qn_stack - 1
                no_of_question = no_of_question + 1
            else:
                print('Warning: LaTeX delimiter not matching')
    if (qn_stack == 0) and (no_of_question > 0):
        text_list = [''] * (2 * no_of_question + 1)
        sub_string = txt
        isequation = [False] * (2 * no_of_question + 1)
        btic = False
        for j in range(0, 2 * no_of_question + 1):
            if btic:
                index = sub_string.find('\\)') + 2
                text_list[j] = sub_string[0: index].replace('<br>', '\n')
                sub_string = sub_string[index: len(sub_string)]
                isequation[j] = True
                btic = False
            else:
                index = sub_string.find('\\(')
                if index == -1:
                    text_list[j] = sub_string
                elif index == 0:
                    text_list[j] = ' '
                else:
                    text_list[j] = sub_string[0: index]
                sub_string = sub_string[index: len(sub_string)]
                btic = True
        return ' '.join(text_list)
    elif (qn_stack != 0) and (no_of_question == 0):
        print('Warning: LaTeX Equation Delimiter Not Matching')
        logging.warning('LaTeX Equation Delimiter Not Matching')
    return txt


def process_cell_text(txt: string, image_name: string):
    txt1 = check_and_break_lines_with_br(txt)
    if image_name != ' ':
        text = "<p>" + txt1 + " </p>, <br> <img src=\"@@PLUGINFILE@@/" + image_name + "\" alt=\" \" " + \
               "role=\"presentation\">" + " <br></p>"
    else:
        # text = "<![CDATA[ <p>" + txt + " <br></p>]]>"
        text = "<p>" + txt1 + "<br> </p>"
    return text


def create_question_name(root, qname):
    nd_question = root.createElement('question')
    nd_question.setAttribute('type', 'multichoice')
    nd_qn_name = root.createElement('name')
    nd_question.appendChild(nd_qn_name)
    nd_qn_txt = root.createElement('text')
    nd_qn_name.appendChild(nd_qn_txt)
    qn_name_txt = root.createTextNode(str(qname))
    nd_qn_txt.appendChild(qn_name_txt)
    nd_shuffle = root.createElement('shuffleanswers')
    nd_question.appendChild(nd_shuffle)
    shuffle_text = root.createTextNode('1')
    nd_shuffle.appendChild(shuffle_text)
    return nd_question


def create_question_text(root, qtext, img_name):
    nd_question = root.createElement('questiontext')
    nd_question.setAttribute('format', 'html')
    nd_qn_txt = root.createElement('text')
    nd_question.appendChild(nd_qn_txt)
    qtext_edited = process_cell_text(qtext, img_name)
    qn_txt = root.createTextNode(qtext_edited)
    nd_qn_txt.appendChild(qn_txt)
    return nd_question


def add_image_base64(root, imageloader, ref_coordinate):
    image = imageloader.get(str(ref_coordinate))
    output_buffer = BytesIO()
    image.save(output_buffer, format='png')  # Save to buffer
    data_string = output_buffer.getvalue()
    img_string = base64.b64encode(data_string)
    nd_string = root.createTextNode(img_string.decode('utf-8'))
    return nd_string


def create_answer_text(root, answer, fraction, image_name):
    nd_answer = root.createElement('answer')
    nd_answer.setAttribute('fraction', fraction)
    nd_answer.setAttribute('format', "html")
    nd_ans_txt = root.createElement('text')
    nd_answer.appendChild(nd_ans_txt)
    answer_text_edited = process_cell_text(answer, image_name)
    ans_txt = root.createTextNode(answer_text_edited)
    nd_ans_txt.appendChild(ans_txt)
    return nd_answer


def create_image_nodes(root, imageloader, ref_coordinate, img_name):
    nd_image = root.createElement('file')
    nd_image.setAttribute('name', img_name)
    nd_image.setAttribute('path', '/')
    nd_image.setAttribute('encoding', 'base64')
    node_image_data = add_image_base64(root, imageloader, ref_coordinate)
    nd_image.appendChild(node_image_data)
    return nd_image


def create_mcq_question(root, imageloader, ref_cell, bimage_present, qname, qtext, c1, c2, c3, c4, soln):
    qn = create_question_name(root, qname)
    image_name = ' '
    if bimage_present[0] == 1:
        image_name = qname + '_qn' + '.png'
    qntxt = create_question_text(root, qtext, image_name)
    if bimage_present[0] == 1:
        ref_coordinate = ref_cell.offset(1, 1).coordinate
        nimg = create_image_nodes(root, imageloader, ref_coordinate, image_name)
        qntxt.appendChild(nimg)
    qn.appendChild(qntxt)

    answers = ["0", "0", "0", "0"]
    answers[soln - 1] = "100"

    image_name = ' '
    if bimage_present[1] == 1:
        image_name = qname + '_ans1' + '.png'
    ans1 = create_answer_text(root, c1, answers[0], image_name)
    if bimage_present[1] == 1:
        ref_coordinate = ref_cell.offset(1, 2).coordinate
        nimg = create_image_nodes(root, imageloader, ref_coordinate, image_name)
        ans1.appendChild(nimg)

    image_name = ' '
    if bimage_present[2] == 1:
        image_name = qname + '_ans2' + '.png'
    ans2 = create_answer_text(root, c2, answers[1], image_name)
    if bimage_present[2] == 1:
        ref_coordinate = ref_cell.offset(1, 3).coordinate
        nimg = create_image_nodes(root, imageloader, ref_coordinate, image_name)
        ans2.appendChild(nimg)

    image_name = ' '
    if bimage_present[3] == 1:
        image_name = qname + '_ans3' + '.png'
    ans3 = create_answer_text(root, c3, answers[2], image_name)
    if bimage_present[3] == 1:
        ref_coordinate = ref_cell.offset(1, 4).coordinate
        nimg = create_image_nodes(root, imageloader, ref_coordinate, image_name)
        ans3.appendChild(nimg)

    image_name = ' '
    if bimage_present[3] == 1:
        image_name = qname + '_ans4' + '.png'
    ans4 = create_answer_text(root, c4, answers[3], image_name)
    if bimage_present[3] == 1:
        ref_coordinate = ref_cell.offset(1, 5).coordinate
        nimg = create_image_nodes(root, imageloader, ref_coordinate, image_name)
        ans4.appendChild(nimg)

    qn.appendChild(ans1)
    qn.appendChild(ans2)
    qn.appendChild(ans3)
    qn.appendChild(ans4)
    return qn


def create_question_category(root, categoryname):
    nd_question = root.createElement('question')
    nd_question.setAttribute('type', 'category')
    nd_qn_category = root.createElement('category')
    nd_question.appendChild(nd_qn_category)
    nd_qn_txt = root.createElement('text')
    nd_qn_category.appendChild(nd_qn_txt)
    qn_text = root.createTextNode('$course$/' + str(categoryname))
    nd_qn_txt.appendChild(qn_text)
    return nd_question


def is_valid_question(start_cell, b_image_present):
    if (start_cell.offset(0, 1).value is None) and (b_image_present[0] == 0):
        logging.error("Question not found for " + start_cell.value)
        print("Question not found for " + start_cell.value)
        return False
    if (start_cell.offset(0, 2).value is None) and (b_image_present[1] == 0):
        logging.error("Answer 1 not found for " + start_cell.value)
        print("Answer 1 not found for " + start_cell.value)
        return False
    if (start_cell.offset(0, 3).value is None) and (b_image_present[2] == 0):
        logging.error("Answer 2 not found for " + start_cell.value)
        print("Answer 2 not found for " + start_cell.value)
        return False
    if (start_cell.offset(0, 4).value is None) and (b_image_present[3] == 0):
        logging.error("Answer 3 not found for " + start_cell.value)
        print("Answer 3 not found for " + start_cell.value)
        return False
    if (start_cell.offset(0, 5).value is None) and (b_image_present[3] == 0):
        logging.error("Answer 4 not found for " + start_cell.value)
        print("Answer 3 not found for " + start_cell.value)
        return False
    if start_cell.offset(0, 6).value is None:
        logging.error("Solution not found for " + start_cell.value)
        print("Solution not found for " + start_cell.value)
        return False
    return True


def check_if_image_present(worksheet, ref_coordinate, imageloader):
    m_base_cell = worksheet[ref_coordinate]
    b_image_present = [0, 0, 0, 0, 0]
    for k in range(1, 5):
        if imageloader.image_in(m_base_cell.offset(1, k).coordinate):
            b_image_present[k - 1] = 1
    return b_image_present


if __name__ == '__main__':
    save_path_file = ' '
    wb = ' '
    # Process the arguments
    args = sys.argv
    if len(args) > 1:
        excel_file = args[1]
        if len(args) > 2:
            save_path_file = args[2]
    else:
        save_path_file = "Quiz.xml"
        excel_file = 'Quiz.xlsx'
    logging.basicConfig(filename='conversion.log', level=logging.INFO,
                        format='%(asctime)s %(levelname)-8s %(message)s',  datefmt='%d/%m/%y %H:%M')
    if not path.exists(excel_file):
        print('Input Excel file: ' + excel_file + ' is not found. Exiting.')
        logging.error('Input Excel file: ' + excel_file + ' is not found. Exiting.')
        exit()
    else:
        # load the workbook
        try:
            wb = load_workbook(str(excel_file))
        except:
            print('Input Excel file: ' + excel_file + ' can not be opened. Exiting.')
            logging.error('Input Excel file: ' + excel_file + ' can not be opened. Exiting.')
            exit()
        finally:
            print('Excel File is found.')

    node_root = minidom.Document()
    node_quiz = node_root.createElement('quiz')
    node_root.appendChild(node_quiz)
    sheet_count = -1
    sheet_names = wb.sheetnames
    for ws in wb.worksheets:
        sheet_count = sheet_count + 1
        row_count = ws.max_row
        base_cell = ws['B1']
        if row_count <= 1:
            print('Empty Sheet. Skipping the sheet: ' + sheet_names[sheet_count])
            logging.warning('Empty Sheet. Skipping the sheet: ' + sheet_names[sheet_count])
            continue
        else:
            print('Processing Sheet: ' + sheet_names[sheet_count])
            logging.info('Processing Sheet: ' + sheet_names[sheet_count])
            image_loader = SheetImageLoader(ws)  # loads the image in the worksheet
            node_quiz_category = create_question_category(node_root, sheet_names[sheet_count])
            node_quiz.appendChild(node_quiz_category)
            for i in range(1, row_count):
                if base_cell.offset(i, 0).value is None:
                    continue
                question_name = str(base_cell.offset(i, 0).value)
                # first element is for question, remaining four are for choices 1 to 4
                bImagePresent = check_if_image_present(ws, base_cell.offset(i, 0).coordinate, image_loader)
                if not is_valid_question(base_cell.offset(i, 0), bImagePresent):
                    print('Skipping Question: '+question_name)
                    logging.error('Skipping Question: '+question_name)
                    continue

                question_txt = str(base_cell.offset(i, 1).value)
                # check whether image exists in the cell in the excel file
                # question_image_name = None
                ans1_txt = str(base_cell.offset(i, 2).value)
                ans2_txt = str(base_cell.offset(i, 3).value)
                ans3_txt = str(base_cell.offset(i, 4).value)
                ans4_txt = str(base_cell.offset(i, 5).value)
                solution = int(base_cell.offset(i, 6).value)
                node_question = create_mcq_question(node_root, image_loader, base_cell.offset(i, 0), bImagePresent,
                                                    question_name, question_txt, ans1_txt, ans2_txt, ans3_txt, ans4_txt,
                                                    solution)
                node_quiz.appendChild(node_question)
                print('Question ' + question_name + ' is added.')
                logging.info('Question ' + question_name + ' is added.')
                # if bImagePresent == 1:
                #     if path.exists(question_image_name):
                #         os.remove(question_image_name)
                #         bImagePresent = 0

    # Convert the xml structure to string to be written to XML
    xml_str = node_root.toprettyxml(indent="\t")
    # xml_str = node_root.toxml()
    with open(save_path_file, "w") as f:
        print('Writing to file: ' + save_path_file)
        f.write(xml_str)
    print('XML Generated Successfully!')
    # print('Enter a key to continue...')
    logging.info('Conversion Completed')
    # dummy=input()
